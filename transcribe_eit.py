#!/usr/bin/env python3
"""
AutoEIT Audio Transcription Script
===================================
GSoC 2026 - Test I: Audio file transcription

This script uses OpenAI Whisper to transcribe Spanish EIT (Elicited Imitation Task)
audio files. It processes 4 participant audio files, segments transcriptions into
30 sentences each, and writes results to an Excel file matching the provided template.

Approach:
---------
1. Use Whisper "medium" model for Spanish - balances accuracy and speed
2. Transcribe with word-level timestamps to enable precise sentence segmentation
3. Use target sentences as prompts to guide Whisper alignment (initial_prompt)
4. Preserve all disfluencies (um, uh, false starts, repetitions)
5. Only correct obvious ASR errors, NOT participant grammar/vocabulary errors

Author: GSoC 2026 Applicant
"""

import whisper
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import re
import json
import time
from copy import copy

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AUDIO_DIR = os.path.join(BASE_DIR, "AutoEIT Test Files", "Sample Audio Files and Transcriptions")
TEMPLATE_FILE = os.path.join(AUDIO_DIR, "AutoEIT Sample Audio for Transcribing.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "AutoEIT Sample Audio for Transcribing_COMPLETED.xlsx")

# Participant configurations
# Each entry: (sheet_name, audio_file, skip_seconds)
PARTICIPANTS = [
    ("38010-2A", "038010_EIT-2A.mp3", 150),   # ~2:30 skip
    ("38011-1A", "038011_EIT-1A.mp3", 150),   # ~2:30 skip
    ("38012-2A", "038012_EIT-2A.mp3", 720),   # 12:00 skip (special note)
    ("38015-1A", "038015_EIT-1A.mp3", 150),   # ~2:30 skip
]

# The 30 target/stimulus sentences (EIT Version A) - used as context for Whisper
TARGET_SENTENCES = [
    "Quiero cortarme el pelo",
    "El libro está en la mesa",
    "El carro lo tiene Pedro",
    "El se ducha cada mañana",
    "¿Qué dice usted que va a hacer hoy?",
    "Dudo que sepa manejar muy bien",
    "Las calles de esta ciudad son muy anchas",
    "Puede que llueva mañana todo el día",
    "Las casas son muy bonitas pero caras",
    "Me gustan las películas que acaban bien",
    "El chico con el que yo salgo es español",
    "Después de cenar me fui a dormir tranquilo",
    "Quiero una casa en la que vivan mis animales",
    "A nosotros nos fascinan las fiestas grandiosas",
    "Ella sólo bebe cerveza y no come nada",
    "Me gustaría que el precio de las casas bajara",
    "Cruza a la derecha y después sigue todo recto",
    "Ella ha terminado de pintar su apartamento",
    "Me gustaría que empezara a hacer más calor pronto",
    "El niño al que se le murió el gato está triste",
    "Una amiga mía cuida a los niños de mi vecino",
    "El gato que era negro fue perseguido por el perro",
    "Antes de poder salir él tiene que limpiar su cuarto",
    "La cantidad de personas que fuman ha disminuido",
    "Después de llegar a casa del trabajo tomé la cena",
    "El ladrón al que atrapó la policía era famoso",
    "Le pedí a un amigo que me ayudara con la tarea",
    "El examen no fue tan difícil como me habían dicho",
    "¿Serías tan amable de darme el libro que está en la mesa?",
    "Hay mucha gente que no toma nada para el desayuno",
]

# Whisper model size - "medium" for good Spanish accuracy on CPU
MODEL_SIZE = "medium"


def load_whisper_model():
    """Load the Whisper model."""
    print(f"Loading Whisper '{MODEL_SIZE}' model...")
    model = whisper.load_model(MODEL_SIZE)
    print("Model loaded successfully.")
    return model


def transcribe_audio(model, audio_path, skip_seconds=150):
    """
    Transcribe a full EIT audio file using Whisper.
    
    Parameters:
    -----------
    model : whisper.Whisper
        Loaded Whisper model
    audio_path : str
        Path to the MP3 audio file
    skip_seconds : int
        Number of seconds to skip from the beginning (instructions/practice)
    
    Returns:
    --------
    dict : Whisper transcription result with segments
    """
    print(f"\n{'='*60}")
    print(f"Transcribing: {os.path.basename(audio_path)}")
    print(f"Skipping first {skip_seconds} seconds (instructions/practice)")
    print(f"{'='*60}")
    
    # Load audio
    audio = whisper.load_audio(audio_path)
    sample_rate = 16000  # Whisper uses 16kHz
    
    # Skip the instruction/practice portion
    start_sample = skip_seconds * sample_rate
    if start_sample < len(audio):
        audio = audio[start_sample:]
    
    # Create context prompt with target sentences to help Whisper
    # This helps with Spanish-specific vocabulary and keeps output in Spanish
    context = "Transcripción de un examen oral de español. El participante repite oraciones en español. " + \
              " ".join(TARGET_SENTENCES[:5])
    
    start_time = time.time()
    
    # Transcribe with Whisper
    result = model.transcribe(
        audio,
        language="es",                    # Force Spanish
        task="transcribe",                # Transcription (not translation)
        initial_prompt=context,           # Context to guide the model
        word_timestamps=True,             # Get word-level timing
        condition_on_previous_text=True,  # Maintain coherence
        verbose=False,                    # Suppress per-segment output
        no_speech_threshold=0.6,          # Detect silence/pauses
        compression_ratio_threshold=2.4,  # Filter hallucinations
        logprob_threshold=-1.0,           # Accept lower-confidence text
        temperature=(0.0, 0.2, 0.4),      # Try multiple temperatures
    )

    elapsed = time.time() - start_time
    print(f"Transcription completed in {elapsed:.1f} seconds")
    print(f"Total segments found: {len(result['segments'])}")
    
    return result


def segment_into_sentences(result, num_expected=30):
    """
    Segment Whisper output into individual sentences.
    
    The EIT audio has a pattern: stimulus plays -> pause -> participant repeats.
    We need to extract only the participant's repetitions (not the stimulus playback).
    
    Strategy:
    - Use Whisper's segment boundaries (it naturally segments on pauses)
    - Group nearby segments that form a single utterance
    - Map to the 30 expected sentences based on timing order
    
    Parameters:
    -----------
    result : dict
        Whisper transcription result
    num_expected : int
        Number of expected sentences (30)
    
    Returns:
    --------
    list : List of transcribed sentences
    """
    segments = result.get("segments", [])
    
    if not segments:
        print("WARNING: No segments found in transcription!")
        return ["[no speech detected]"] * num_expected
    
    # Collect all segments with their text and timing
    all_segments = []
    for seg in segments:
        text = seg["text"].strip()
        if text and text not in ["...", ".", ""]:
            all_segments.append({
                "start": seg["start"],
                "end": seg["end"],
                "text": text,
                "no_speech_prob": seg.get("no_speech_prob", 0),
            })
    
    print(f"  Meaningful segments: {len(all_segments)}")
    
    # In EIT, there's a pattern: stimulus (recorded voice) -> gap -> response (participant)
    # The gaps between stimulus+response pairs are typically longer than within-pair gaps
    # We need to group segments into utterance groups
    
    if len(all_segments) == 0:
        return ["[no speech detected]"] * num_expected
    
    # Calculate inter-segment gaps
    gaps = []
    for i in range(1, len(all_segments)):
        gap = all_segments[i]["start"] - all_segments[i-1]["end"]
        gaps.append(gap)
    
    # In EIT, the audio has pairs: (stimulus plays, then participant repeats)
    # Each pair is separated by a longer gap
    # We group segments that are close together (< threshold)
    
    # Determine grouping threshold
    # Typical pattern: ~3-8 second gaps between items, ~0.5-2 second gaps within
    GROUP_GAP_THRESHOLD = 3.0  # seconds - segments closer than this are grouped
    
    # Group segments into utterances
    utterance_groups = []
    current_group = [all_segments[0]]
    
    for i in range(1, len(all_segments)):
        gap = all_segments[i]["start"] - current_group[-1]["end"]
        if gap > GROUP_GAP_THRESHOLD:
            utterance_groups.append(current_group)
            current_group = [all_segments[i]]
        else:
            current_group.append(all_segments[i])
    utterance_groups.append(current_group)
    
    print(f"  Utterance groups formed: {len(utterance_groups)}")
    
    # Each group contains either stimulus OR response OR both
    # In EIT, the stimulus is played from a recording and the participant repeats
    # Both may be captured by Whisper
    
    # Build final sentences from groups
    # If we have roughly 2x expected groups, every other one is likely the response
    # If we have roughly 1x expected, each group is the response
    
    sentences = []
    for group in utterance_groups:
        combined = " ".join(seg["text"] for seg in group)
        # Clean up but preserve disfluencies
        combined = clean_transcription(combined)
        sentences.append(combined)
    
    # Now we need exactly 30 sentences
    # Strategy: if we have too many, we need to identify stimulus vs response
    # If too few, some responses might have been merged or missed
    
    if len(sentences) > num_expected * 1.5:
        # Likely have both stimulus and response - try to separate
        # The stimulus will closely match target sentences
        # The response may differ (learner errors)
        print(f"  Too many groups ({len(sentences)}), filtering stimulus playback...")
        sentences = filter_stimulus_from_responses(sentences, TARGET_SENTENCES, num_expected)
    elif len(sentences) < num_expected:
        # Too few - might need to split some groups or pad
        print(f"  Too few groups ({len(sentences)}), attempting to split...")
        sentences = adjust_sentence_count(sentences, num_expected)
    
    # Ensure exactly 30 sentences
    if len(sentences) > num_expected:
        sentences = sentences[:num_expected]
    while len(sentences) < num_expected:
        sentences.append("[inaudible]")
    
    return sentences


def clean_transcription(text):
    """
    Clean up ASR artifacts while preserving genuine disfluencies.
    
    Rules:
    - Keep participant's grammar errors (NOT corrections)
    - Keep disfluencies: uh, um, eh, ah, false starts, repetitions
    - Remove obvious Whisper artifacts: repeated hallucinated text
    - Fix obvious ASR glitches: random characters, encoding issues
    """
    # Remove leading/trailing whitespace
    text = text.strip()
    
    # Remove obvious Whisper hallucination patterns
    # (e.g., same phrase repeated many times)
    words = text.split()
    if len(words) > 4:
        # Check for repeating patterns of 3+ words at the end
        for pattern_len in range(3, len(words)//2 + 1):
            pattern = words[-pattern_len:]
            prev = words[-2*pattern_len:-pattern_len]
            if pattern == prev:
                text = " ".join(words[:-pattern_len])
                break
    
    # Remove music/sound markers that Whisper sometimes adds
    text = re.sub(r'\[.*?\]', '', text)
    text = re.sub(r'\(.*?música.*?\)', '', text, flags=re.IGNORECASE)
    
    # Fix double spaces
    text = re.sub(r'\s+', ' ', text).strip()
    
    # Remove leading/trailing punctuation artifacts
    text = text.strip('.')
    text = text.strip()
    
    return text


def filter_stimulus_from_responses(sentences, targets, num_expected):
    """
    When Whisper captures both stimulus and response, filter out stimulus playback.
    
    Uses similarity matching: if a segment closely matches a target sentence,
    it's likely the stimulus playback, not the participant's response.
    """
    from difflib import SequenceMatcher
    
    scored = []
    for sent in sentences:
        # Check similarity to each target
        max_sim = 0
        for target in targets:
            sim = SequenceMatcher(None, sent.lower(), target.lower()).ratio()
            max_sim = max(max_sim, sim)
        scored.append((sent, max_sim))
    
    # If similarity is very high (>0.9), it might be the stimulus
    # But we need to be careful - good learners might produce very similar text
    # Strategy: if we have pairs, take every other one (the response follows stimulus)
    
    if len(sentences) >= num_expected * 1.8:
        # Clearly have stimulus+response pairs
        # Take alternating entries (response comes after stimulus)
        responses = []
        for i in range(0, len(sentences), 2):
            if i + 1 < len(sentences):
                # Compare both to target - the one less similar is likely the response
                sim_first = scored[i][1]
                sim_second = scored[i+1][1]
                if sim_first > sim_second:
                    responses.append(sentences[i+1])  # Second is response
                else:
                    responses.append(sentences[i])     # First is response
            else:
                responses.append(sentences[i])
        return responses[:num_expected]
    else:
        # Moderate excess - just take most likely responses
        # Sort by position, remove highest-similarity entries
        return sentences[:num_expected]


def adjust_sentence_count(sentences, num_expected):
    """
    Handle case where we have too few sentences.
    Attempt to split long segments or pad with markers.
    """
    if len(sentences) >= num_expected:
        return sentences[:num_expected]
    
    # Try splitting very long sentences that might contain multiple utterances
    expanded = []
    for sent in sentences:
        # If a sentence is unusually long, try to split on major pauses
        words = sent.split()
        if len(words) > 20:
            # Split roughly in half at a natural point
            mid = len(words) // 2
            expanded.append(" ".join(words[:mid]))
            expanded.append(" ".join(words[mid:]))
        else:
            expanded.append(sent)
    
    return expanded


def write_results_to_excel(all_transcriptions):
    """
    Write transcription results to Excel file based on template format.
    
    Parameters:
    -----------
    all_transcriptions : dict
        {sheet_name: [list of 30 transcribed sentences]}
    """
    print(f"\nWriting results to: {OUTPUT_FILE}")
    
    # Load template
    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    
    for sheet_name, transcriptions in all_transcriptions.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Writing {len(transcriptions)} transcriptions to sheet '{sheet_name}'")
            
            for i, text in enumerate(transcriptions):
                # Transcription goes in column C (column 3), rows 2-31
                row = i + 2  # Row 1 is header
                ws.cell(row=row, column=3, value=text)
                # Style the cell
                ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        else:
            print(f"  WARNING: Sheet '{sheet_name}' not found in template!")
    
    wb.save(OUTPUT_FILE)
    print(f"Results saved to: {OUTPUT_FILE}")


def save_raw_transcriptions(all_results):
    """Save raw Whisper output for debugging/analysis."""
    raw_dir = os.path.join(BASE_DIR, "raw_transcriptions")
    os.makedirs(raw_dir, exist_ok=True)
    
    for sheet_name, result in all_results.items():
        out_file = os.path.join(raw_dir, f"{sheet_name}_raw.json")
        # Convert to serializable format
        serializable = {
            "text": result["text"],
            "segments": [
                {
                    "start": seg["start"],
                    "end": seg["end"],
                    "text": seg["text"],
                    "no_speech_prob": seg.get("no_speech_prob", 0),
                }
                for seg in result.get("segments", [])
            ]
        }
        with open(out_file, 'w', encoding='utf-8') as f:
            json.dump(serializable, f, ensure_ascii=False, indent=2)
        print(f"  Raw output saved: {out_file}")


def main():
    """Main execution flow."""
    print("=" * 60)
    print("AutoEIT Audio Transcription - GSoC 2026 Test I")
    print("=" * 60)
    print(f"\nModel: Whisper {MODEL_SIZE}")
    print(f"Audio directory: {AUDIO_DIR}")
    print(f"Output file: {OUTPUT_FILE}")
    print(f"Participants: {len(PARTICIPANTS)}")
    
    # Load model
    model = load_whisper_model()
    
    all_transcriptions = {}  # sheet_name -> [30 sentences]
    all_raw_results = {}     # sheet_name -> whisper result
    
    for sheet_name, audio_file, skip_secs in PARTICIPANTS:
        audio_path = os.path.join(AUDIO_DIR, audio_file)
        
        if not os.path.exists(audio_path):
            print(f"\nERROR: Audio file not found: {audio_path}")
            all_transcriptions[sheet_name] = ["[file not found]"] * 30
            continue
        
        # Transcribe
        result = transcribe_audio(model, audio_path, skip_seconds=skip_secs)
        all_raw_results[sheet_name] = result
        
        # Segment into 30 sentences
        sentences = segment_into_sentences(result, num_expected=30)
        all_transcriptions[sheet_name] = sentences
        
        # Print preview
        print(f"\n  Preview of transcriptions for {sheet_name}:")
        for i, sent in enumerate(sentences[:5]):
            print(f"    {i+1}. {sent}")
        print(f"    ... ({len(sentences)} total)")
    
    # Save raw transcriptions for debugging
    print("\nSaving raw transcriptions...")
    save_raw_transcriptions(all_raw_results)
    
    # Write to Excel
    write_results_to_excel(all_transcriptions)
    
    # Summary
    print("\n" + "=" * 60)
    print("TRANSCRIPTION COMPLETE")
    print("=" * 60)
    for sheet_name, sents in all_transcriptions.items():
        filled = sum(1 for s in sents if s not in ["[inaudible]", "[no speech detected]", "[file not found]"])
        print(f"  {sheet_name}: {filled}/30 sentences transcribed")
    print(f"\nOutput: {OUTPUT_FILE}")
    print(f"Raw data: {os.path.join(BASE_DIR, 'raw_transcriptions')}")


if __name__ == "__main__":
    main()
